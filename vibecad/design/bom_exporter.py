"""
BOM (Bill of Materials) Exporter for VibeCAD.

Supports multiple output formats:
- CSV (generic, JLCPCB, LCSC, Mouser, DigiKey)
- Excel (XLSX)
- JSON
- HTML

All exports require user approval with preview.
"""

import csv
import json
import logging
import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from io import StringIO
from typing import Optional, List, Dict, Any, Callable

logger = logging.getLogger(__name__)


class BOMFormat(Enum):
    """Supported BOM export formats."""
    CSV_GENERIC = "csv_generic"
    CSV_JLCPCB = "csv_jlcpcb"
    CSV_LCSC = "csv_lcsc"
    CSV_MOUSER = "csv_mouser"
    CSV_DIGIKEY = "csv_digikey"
    EXCEL = "excel"
    JSON = "json"
    HTML = "html"


@dataclass
class BOMEntry:
    """A single BOM entry (grouped by value/footprint)."""
    references: List[str]  # e.g., ["R1", "R2", "R3"]
    value: str  # e.g., "10k"
    footprint: str  # e.g., "0603"
    quantity: int
    
    # Optional manufacturer data
    manufacturer: Optional[str] = None
    mpn: Optional[str] = None  # Manufacturer Part Number
    
    # Supplier data
    lcsc_part: Optional[str] = None  # For JLCPCB
    mouser_part: Optional[str] = None
    digikey_part: Optional[str] = None
    
    # Additional fields
    description: Optional[str] = None
    datasheet: Optional[str] = None
    dnp: bool = False  # Do Not Place
    
    # Custom properties from schematic
    properties: Dict[str, str] = field(default_factory=dict)
    
    def references_str(self, delimiter: str = ", ") -> str:
        """Get references as a delimited string."""
        return delimiter.join(sorted(self.references, key=self._ref_sort_key))
    
    @staticmethod
    def _ref_sort_key(ref: str):
        """Sort references naturally (R1, R2, R10 not R1, R10, R2)."""
        import re
        match = re.match(r'([A-Za-z]+)(\d+)', ref)
        if match:
            return (match.group(1), int(match.group(2)))
        return (ref, 0)


@dataclass
class BOMExportRequest:
    """Request to export a BOM."""
    entries: List[BOMEntry]
    format: BOMFormat = BOMFormat.CSV_GENERIC
    output_path: Optional[str] = None
    
    # Options
    include_dnp: bool = False
    group_by: str = "value_footprint"  # or "value", "mpn"
    sort_by: str = "reference"  # or "value", "quantity"
    
    # Metadata
    project_name: Optional[str] = None
    revision: Optional[str] = None
    date: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'format': self.format.value,
            'entry_count': len(self.entries),
            'total_components': sum(e.quantity for e in self.entries),
            'include_dnp': self.include_dnp,
            'project_name': self.project_name,
            'date': self.date,
        }


@dataclass
class BOMExportResult:
    """Result of a BOM export operation."""
    success: bool
    message: str
    output_path: Optional[str] = None
    preview_data: Optional[str] = None
    error: Optional[str] = None
    total_unique_parts: int = 0
    total_components: int = 0


class BOMExporter:
    """Exports BOM in various formats.
    
    Core principles:
    - User previews BOM before export
    - Supports multiple manufacturer formats
    - Can extract component data from PCB or schematic
    """
    
    # Format-specific column mappings
    FORMAT_COLUMNS = {
        BOMFormat.CSV_GENERIC: [
            "Reference", "Value", "Footprint", "Quantity", "Manufacturer", "MPN", "Description"
        ],
        BOMFormat.CSV_JLCPCB: [
            "Comment", "Designator", "Footprint", "LCSC Part #"
        ],
        BOMFormat.CSV_LCSC: [
            "Quantity", "Manufacturer Part Number", "Manufacturer", "Description", "Customer Reference"
        ],
        BOMFormat.CSV_MOUSER: [
            "Mouser Part Number", "Manufacturer Part Number", "Manufacturer", "Description", "Quantity"
        ],
        BOMFormat.CSV_DIGIKEY: [
            "Quantity", "Part Number", "Manufacturer Part Number", "Manufacturer Name", "Description"
        ],
    }
    
    def __init__(self):
        self._progress_callback: Optional[Callable] = None
    
    def set_progress_callback(self, callback: Callable):
        """Set callback for export progress updates."""
        self._progress_callback = callback
    
    def extract_from_pcb(self, pcb_data: Any) -> List[BOMEntry]:
        """Extract BOM entries from parsed PCB data.
        
        Args:
            pcb_data: Parsed PCB data from PCBParser
        
        Returns:
            List of BOMEntry objects
        """
        # Group components by value and footprint
        groups: Dict[tuple, List[str]] = {}
        
        footprints = getattr(pcb_data, 'footprints', [])
        
        for fp in footprints:
            ref = getattr(fp, 'reference', '?')
            value = getattr(fp, 'value', '')
            
            # Get footprint name from lib_id
            fp_name = getattr(fp, 'lib_id', '') or ''
            if ':' in fp_name:
                fp_name = fp_name.split(':')[-1]
            
            # Check for DNP
            dnp = getattr(fp, 'dnp', False)
            
            # Create grouping key
            key = (value, fp_name, dnp)
            
            if key not in groups:
                groups[key] = []
            groups[key].append(ref)
        
        # Convert groups to BOM entries
        entries = []
        for (value, footprint, dnp), refs in groups.items():
            # Try to extract properties from first footprint
            entry = BOMEntry(
                references=refs,
                value=value,
                footprint=footprint,
                quantity=len(refs),
                dnp=dnp,
            )
            
            # TODO: Extract manufacturer data from schematic properties
            
            entries.append(entry)
        
        return entries
    
    def extract_from_schematic(self, schematic_data: Any) -> List[BOMEntry]:
        """Extract BOM entries from parsed schematic data.
        
        Args:
            schematic_data: Parsed schematic data
        
        Returns:
            List of BOMEntry objects
        """
        # Group components by value and footprint
        groups: Dict[tuple, Dict[str, Any]] = {}
        
        components = getattr(schematic_data, 'components', [])
        
        for comp in components:
            ref = getattr(comp, 'reference', '?')
            value = getattr(comp, 'value', '')
            footprint = getattr(comp, 'footprint', '')
            
            # Get properties
            props = getattr(comp, 'properties', {})
            mpn = props.get('MPN', props.get('Manufacturer Part Number', ''))
            manufacturer = props.get('Manufacturer', '')
            lcsc = props.get('LCSC', props.get('LCSC Part', ''))
            
            dnp = getattr(comp, 'dnp', False)
            
            # Create grouping key
            key = (value, footprint, mpn, manufacturer)
            
            if key not in groups:
                groups[key] = {
                    'refs': [],
                    'value': value,
                    'footprint': footprint,
                    'mpn': mpn,
                    'manufacturer': manufacturer,
                    'lcsc': lcsc,
                    'dnp': dnp,
                    'props': props,
                }
            groups[key]['refs'].append(ref)
        
        # Convert groups to BOM entries
        entries = []
        for key, data in groups.items():
            entry = BOMEntry(
                references=data['refs'],
                value=data['value'],
                footprint=data['footprint'],
                quantity=len(data['refs']),
                manufacturer=data['manufacturer'] or None,
                mpn=data['mpn'] or None,
                lcsc_part=data['lcsc'] or None,
                dnp=data['dnp'],
                properties=data['props'],
            )
            entries.append(entry)
        
        return entries
    
    def create_preview(self, request: BOMExportRequest) -> str:
        """Create a preview of the BOM export.
        
        Args:
            request: The export request
        
        Returns:
            String preview of the BOM
        """
        lines = [
            "📋 BOM Export Preview",
            "",
            f"Format: {request.format.value}",
            f"Total unique parts: {len(request.entries)}",
            f"Total components: {sum(e.quantity for e in request.entries)}",
            "",
        ]
        
        if request.project_name:
            lines.append(f"Project: {request.project_name}")
        if request.revision:
            lines.append(f"Revision: {request.revision}")
        lines.append(f"Date: {request.date}")
        lines.append("")
        
        # Show first few entries as preview
        lines.append("Preview (first 10 entries):")
        lines.append("-" * 60)
        
        entries = request.entries[:10]
        for entry in entries:
            refs = entry.references_str()
            if len(refs) > 30:
                refs = refs[:27] + "..."
            lines.append(f"  {refs:<30} | {entry.value:<15} | x{entry.quantity}")
        
        if len(request.entries) > 10:
            lines.append(f"  ... and {len(request.entries) - 10} more entries")
        
        lines.append("-" * 60)
        
        if request.output_path:
            lines.append(f"\nOutput file: {request.output_path}")
        
        return "\n".join(lines)
    
    def export(self, request: BOMExportRequest) -> BOMExportResult:
        """Export BOM to the specified format.
        
        Args:
            request: The export request
        
        Returns:
            BOMExportResult with success status
        """
        # Filter DNP if requested
        entries = request.entries
        if not request.include_dnp:
            entries = [e for e in entries if not e.dnp]
        
        # Sort entries
        entries = self._sort_entries(entries, request.sort_by)
        
        try:
            if request.format in (BOMFormat.CSV_GENERIC, BOMFormat.CSV_JLCPCB, 
                                 BOMFormat.CSV_LCSC, BOMFormat.CSV_MOUSER, 
                                 BOMFormat.CSV_DIGIKEY):
                content = self._export_csv(entries, request.format)
            elif request.format == BOMFormat.JSON:
                content = self._export_json(entries, request)
            elif request.format == BOMFormat.HTML:
                content = self._export_html(entries, request)
            elif request.format == BOMFormat.EXCEL:
                return self._export_excel(entries, request)
            else:
                return BOMExportResult(
                    success=False,
                    message=f"Unsupported format: {request.format}",
                    error="Format not implemented",
                )
            
            # Write to file if path specified
            if request.output_path:
                with open(request.output_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                
                return BOMExportResult(
                    success=True,
                    message=f"BOM exported to {request.output_path}",
                    output_path=request.output_path,
                    total_unique_parts=len(entries),
                    total_components=sum(e.quantity for e in entries),
                )
            else:
                return BOMExportResult(
                    success=True,
                    message="BOM generated",
                    preview_data=content,
                    total_unique_parts=len(entries),
                    total_components=sum(e.quantity for e in entries),
                )
                
        except Exception as e:
            logger.exception(f"BOM export failed: {e}")
            return BOMExportResult(
                success=False,
                message=f"Export failed: {e}",
                error=str(e),
            )
    
    def _sort_entries(self, entries: List[BOMEntry], sort_by: str) -> List[BOMEntry]:
        """Sort BOM entries."""
        if sort_by == "reference":
            return sorted(entries, key=lambda e: e._ref_sort_key(e.references[0]) if e.references else ('', 0))
        elif sort_by == "value":
            return sorted(entries, key=lambda e: e.value.lower())
        elif sort_by == "quantity":
            return sorted(entries, key=lambda e: -e.quantity)
        return entries
    
    def _export_csv(self, entries: List[BOMEntry], format: BOMFormat) -> str:
        """Export to CSV format."""
        output = StringIO()
        
        columns = self.FORMAT_COLUMNS.get(format, self.FORMAT_COLUMNS[BOMFormat.CSV_GENERIC])
        
        writer = csv.writer(output)
        writer.writerow(columns)
        
        for entry in entries:
            if format == BOMFormat.CSV_GENERIC:
                row = [
                    entry.references_str(),
                    entry.value,
                    entry.footprint,
                    entry.quantity,
                    entry.manufacturer or "",
                    entry.mpn or "",
                    entry.description or "",
                ]
            elif format == BOMFormat.CSV_JLCPCB:
                row = [
                    entry.value,  # Comment
                    entry.references_str(),  # Designator
                    entry.footprint,  # Footprint
                    entry.lcsc_part or "",  # LCSC Part #
                ]
            elif format == BOMFormat.CSV_LCSC:
                row = [
                    entry.quantity,
                    entry.mpn or "",
                    entry.manufacturer or "",
                    entry.description or entry.value,
                    entry.references_str(),  # Customer Reference
                ]
            elif format == BOMFormat.CSV_MOUSER:
                row = [
                    entry.mouser_part or "",
                    entry.mpn or "",
                    entry.manufacturer or "",
                    entry.description or entry.value,
                    entry.quantity,
                ]
            elif format == BOMFormat.CSV_DIGIKEY:
                row = [
                    entry.quantity,
                    entry.digikey_part or "",
                    entry.mpn or "",
                    entry.manufacturer or "",
                    entry.description or entry.value,
                ]
            else:
                row = [entry.references_str(), entry.value, entry.quantity]
            
            writer.writerow(row)
        
        return output.getvalue()
    
    def _export_json(self, entries: List[BOMEntry], request: BOMExportRequest) -> str:
        """Export to JSON format."""
        data = {
            "project": request.project_name,
            "revision": request.revision,
            "date": request.date,
            "total_unique_parts": len(entries),
            "total_components": sum(e.quantity for e in entries),
            "parts": [],
        }
        
        for entry in entries:
            part = {
                "references": entry.references,
                "value": entry.value,
                "footprint": entry.footprint,
                "quantity": entry.quantity,
            }
            
            if entry.manufacturer:
                part["manufacturer"] = entry.manufacturer
            if entry.mpn:
                part["mpn"] = entry.mpn
            if entry.lcsc_part:
                part["lcsc_part"] = entry.lcsc_part
            if entry.description:
                part["description"] = entry.description
            if entry.dnp:
                part["dnp"] = True
            
            data["parts"].append(part)
        
        return json.dumps(data, indent=2)
    
    def _export_html(self, entries: List[BOMEntry], request: BOMExportRequest) -> str:
        """Export to HTML format."""
        total_qty = sum(e.quantity for e in entries)
        
        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>BOM - {request.project_name or 'Project'}</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 20px; }}
        h1 {{ color: #333; }}
        .meta {{ color: #666; margin-bottom: 20px; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #4CAF50; color: white; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
        tr:hover {{ background-color: #ddd; }}
        .qty {{ text-align: center; }}
        .total {{ font-weight: bold; margin-top: 10px; }}
    </style>
</head>
<body>
    <h1>Bill of Materials</h1>
    <div class="meta">
        {f'<p>Project: {request.project_name}</p>' if request.project_name else ''}
        {f'<p>Revision: {request.revision}</p>' if request.revision else ''}
        <p>Date: {request.date}</p>
        <p>Unique parts: {len(entries)} | Total components: {total_qty}</p>
    </div>
    <table>
        <thead>
            <tr>
                <th>#</th>
                <th>References</th>
                <th>Value</th>
                <th>Footprint</th>
                <th class="qty">Qty</th>
                <th>Manufacturer</th>
                <th>MPN</th>
            </tr>
        </thead>
        <tbody>
"""
        for i, entry in enumerate(entries, 1):
            html += f"""            <tr>
                <td>{i}</td>
                <td>{entry.references_str()}</td>
                <td>{entry.value}</td>
                <td>{entry.footprint}</td>
                <td class="qty">{entry.quantity}</td>
                <td>{entry.manufacturer or ''}</td>
                <td>{entry.mpn or ''}</td>
            </tr>
"""
        
        html += """        </tbody>
    </table>
    <p class="total">Generated by VibeCAD</p>
</body>
</html>
"""
        return html
    
    def _export_excel(self, entries: List[BOMEntry], request: BOMExportRequest) -> BOMExportResult:
        """Export to Excel format."""
        try:
            import openpyxl
        except ImportError:
            return BOMExportResult(
                success=False,
                message="Excel export requires openpyxl package",
                error="Install with: pip install openpyxl",
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        
        # Header row
        headers = ["#", "References", "Value", "Footprint", "Quantity", "Manufacturer", "MPN", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Data rows
        for row_num, entry in enumerate(entries, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=entry.references_str())
            ws.cell(row=row_num, column=3, value=entry.value)
            ws.cell(row=row_num, column=4, value=entry.footprint)
            ws.cell(row=row_num, column=5, value=entry.quantity)
            ws.cell(row=row_num, column=6, value=entry.manufacturer or "")
            ws.cell(row=row_num, column=7, value=entry.mpn or "")
            ws.cell(row=row_num, column=8, value=entry.description or "")
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        if request.output_path:
            wb.save(request.output_path)
            return BOMExportResult(
                success=True,
                message=f"Excel BOM exported to {request.output_path}",
                output_path=request.output_path,
                total_unique_parts=len(entries),
                total_components=sum(e.quantity for e in entries),
            )
        else:
            return BOMExportResult(
                success=False,
                message="Excel export requires output_path",
                error="No output path specified",
            )
